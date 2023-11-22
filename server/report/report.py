import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import Alignment
from datetime import datetime

START_DATA_ROW_INDEX = 7


def writeRowData(sheet: Worksheet, data: list, rowIdx: int = 0):
    allSideBorder = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000'),
    )
    for i in range(len(data)):
        cell = sheet.cell(row=rowIdx, column=i + 1, value=data[i])
        cell.border = allSideBorder
        if (len(data) - 1) == i:
            cell.alignment = Alignment(horizontal='right')


def generateReport(rowDatas: list[list]):
    reportDirPath = os.getcwd()
    reportPath = reportDirPath + '/template.xlsx'
    # get workbook and current worksheet
    # wb = xlsxwriter.Workbook(report_path)
    wb: Workbook = load_workbook(reportPath)
    ws: Worksheet = wb.active

    # handle fill data
    today = datetime.now().date()
    # fill time report
    ws['B4'] = today.strftime('%d/%m/%Y')

    # insert row before write
    ws.insert_rows(START_DATA_ROW_INDEX, len(rowDatas))
    for i in range(len(rowDatas)):
        writeRowData(ws, rowDatas[i], START_DATA_ROW_INDEX + i)
    wb.save('Report_' + today.strftime('%d_%m_%Y') + '.xlsx')


testDatas = [
    ['00001', 'Phan Nguyễn Anh Vinh', '09:00:00', '18:00:00', '9'],
    ['00002', 'Nguyễn Nghĩa', '09:00:00', '18:00:00', '9'],
    ['00003', 'Lê Công Minh Khôi', '09:00:00', '18:00:00', '9'],
]

generateReport(testDatas)
